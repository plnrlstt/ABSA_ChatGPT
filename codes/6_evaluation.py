import openpyxl
import pandas as pd

def read_excel_workbook(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Create a dictionary to store structured data for each sheet
    structured_data = {}
    
    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Ignore the first row and column
        rows = sheet.iter_rows(min_row=2, min_col=2)
        
        # Create a list to store structured data for the current sheet
        sheet_data = []
        
        # Iterate through rows and group values in sublists of four
        for row in rows:
            row_values = []
            for cell in row:
                if cell.value is not None:  # Check if cell value is not None
                    row_values.append(cell.value)
                if len(row_values) == 4:
                    sheet_data.append(row_values)
                    row_values = []
            if row_values:
                sheet_data.append(row_values)
        
        # Store structured data for the current sheet in the dictionary
        structured_data[sheet_name] = sheet_data
    
    return structured_data

# Example usage




def calculate_metrics(result1, result2):
    # Create a list to store metrics for each sheet
    sheet_metrics = []
    
    # Iterate through each sheet in result1 (gold standard)
    for sheet_name, gold_standard in result1.items():
        predictions = result2.get(sheet_name)  # Get corresponding predictions
        if predictions is None:
            continue
        
        # Initialize counts for TP, FP, FN
        tp = fp = fn = 0
        
        # Iterate through each sublist in gold standard and predictions
        for gold_sublist in gold_standard:
            for pred_sublist in predictions:
                # Check for true positives (TP)
                if gold_sublist == pred_sublist:
                    tp += 1
                    break
        
        # False positives (FP) is the number of predictions that were not in the gold standard
        fp = len(predictions) - tp
        
        # False negatives (FN) is the number of gold standard that were not in the predictions
        fn = len(gold_standard) - tp
        
        # Calculate precision and recall
        precision = tp / (tp + fp) if tp + fp > 0 else 0
        recall = tp / (tp + fn) if tp + fn > 0 else 0
        
        # Append metrics for the current sheet to the list
        sheet_metrics.append({
            'Sheet': sheet_name,
            'Precision': precision,
            'Recall': recall,
            'TP': tp,
            'FP': fp,
            'FN': fn
        })
    
    return sheet_metrics

# Example usage
# result1 = {...}  # Replace with your gold standard structured data
# result2 = {...}  # Replace with your prediction structured data
file_path1 = r"C:\Users\kpsst\Desktop\FSem\output_sorted (3).xlsx" # Replace with the path to your Excel file
# file_path2
result2 = read_excel_workbook(file_path1)
# result2 = read_excel_workbook(file_path)
# print(result1)
result1=read_excel_workbook(r"C:\Users\kpsst\Desktop\FSem\reviewsfin_BEPL.xlsm")
metrics = calculate_metrics(result1, result2)

# Convert the list of dictionaries to a DataFrame
metrics_df = pd.DataFrame(metrics)

# Save the DataFrame as a CSV file
metrics_df.to_csv('Results.csv', index=False)

# metrics = calculate_metrics(result1, result2)
# # for sheet_name, metric_data in metrics.items():
#     print(f"Sheet: {sheet_name}")
#     print(f"Precision: {metric_data['Precision']}, Recall: {metric_data['Recall']}")
#     print(f"TP: {metric_data['TP']}, FP: {metric_data['FP']}, FN: {metric_data['FN']}")

print(result2)
