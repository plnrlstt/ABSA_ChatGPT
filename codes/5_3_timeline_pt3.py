import openai
import openpyxl


# Set up OpenAI API key
openai.api_key = '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'


def count_negative_reviews(file_path):
   # Load Excel file with specific sheet name
   wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
   sheet = wb['reviews']


   # Dictionary to store counts of negative reviews for each aspect category
   aspect_counts = {}


   # Get column indices for aspect categories and sentiment polarities
   aspect_category_columns = []
   sentiment_polarity_columns = []


   for col in sheet[1]:
       if col.value and col.value.startswith("aspect_category_"):
           aspect_category_columns.append(col.column)
       elif col.value and col.value.startswith("sentiment_polarity_"):
           sentiment_polarity_columns.append(col.column)


   # Iterate through rows
   for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is header
       for aspect_col, sentiment_col in zip(aspect_category_columns, sentiment_polarity_columns):
           aspect_category = row[aspect_col - 1]
           if aspect_category not in aspect_counts:
               aspect_counts[aspect_category] = 0
           # Check if sentiment is negative
           if row[sentiment_col - 1] == 'negative':
               aspect_counts[aspect_category] += 1


   # Write results to new Excel file
   output_wb = openpyxl.Workbook()
   output_sheet = output_wb.active
   output_sheet.append(['Aspect Category', 'Negative Review Count'])
   for aspect, count in aspect_counts.items():
       output_sheet.append([aspect, count])
  
   output_file_path = 'tabelle_cat_neg.xlsx'
   output_wb.save(output_file_path)


   print(f"Results saved to {output_file_path}")


# Example usage
count_negative_reviews('reviewsfin.xlsm')
